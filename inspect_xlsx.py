from __future__ import annotations

from typing import Iterable

from openpyxl import load_workbook


def _row_values(ws, r: int, max_cols: int) -> list[object]:
    return [ws.cell(r, c).value for c in range(1, min(ws.max_column, max_cols) + 1)]


def inspect(paths: Iterable[str]) -> None:
    for path in paths:
        wb = load_workbook(path, data_only=True)
        print(f"\nFILE {path}")
        print("sheets:", ", ".join(wb.sheetnames))
        for name in wb.sheetnames:
            ws = wb[name]
            headers = _row_values(ws, 1, 20)
            print(f"  - sheet={name} rows={ws.max_row} cols={ws.max_column}")
            print(f"    headers={headers}")
            shown = 0
            for r in range(1, min(ws.max_row, 30) + 1):
                row = _row_values(ws, r, 20)
                if any(v not in (None, "") for v in row):
                    print(f"    r{r}={row}")
                    shown += 1
                if shown >= 8:
                    break


if __name__ == "__main__":
    inspect(
        [
            "每周排课操作数（近一年）_2026_03_13.xlsx",
            "每周课表查询次数（近一年）_2026_03_13.xlsx",
            "每周德育分数录入数（近_13_个月）_2026_03_13.xlsx",
            "考试分数录入数_2026_03_13.xlsx",
        ]
    )

