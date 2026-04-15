from __future__ import annotations

import hashlib
import logging
import os
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable, Sequence

import matplotlib

# Important: must set backend before importing pyplot.
matplotlib.use("Agg")
import matplotlib.pyplot as plt  # noqa: E402
from openpyxl import load_workbook  # noqa: E402


DEFAULT_FONT_SANS = [
    "PingFang SC",
    "Hiragino Sans GB",
    "Heiti SC",
    "Arial Unicode MS",
    "DejaVu Sans",
]


def init_fonts() -> None:
    """Init matplotlib font settings for Chinese labels."""
    plt.rcParams["font.sans-serif"] = list(DEFAULT_FONT_SANS)
    plt.rcParams["axes.unicode_minus"] = False


def key_from_filename(filename: str | os.PathLike[str]) -> str:
    """
    Match rule:
    - Take the base name (without extension)
    - If '_' exists, take the substring before the first '_'
    - Otherwise, return the whole base name
    """
    p = Path(filename)
    stem = p.stem
    if "_" not in stem:
        return stem
    return stem.split("_", 1)[0]


def scan_xlsx_by_key(data_dir: str | os.PathLike[str], required_key: str) -> list[Path]:
    data_path = Path(data_dir)
    matches: list[Path] = []
    for p in sorted(data_path.glob("*.xlsx")):
        if key_from_filename(p) == required_key:
            matches.append(p)
    return matches


def parse_date(v: object) -> str:
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v).strip()
    if "T" in s:
        s = s.split("T", 1)[0]
    return s


def load_result_sheet(xlsx_path: str | os.PathLike[str]):
    wb = load_workbook(xlsx_path, data_only=True)
    if "result" in wb.sheetnames:
        ws = wb["result"]
    else:
        ws = wb[wb.sheetnames[0]]
    return wb, ws


def stable_slug(s: str, prefix: str = "auto") -> str:
    """
    Create a stable ASCII slug for filenames (Windows-friendly).
    We keep it deterministic per key.
    """
    h = hashlib.sha1(s.encode("utf-8")).hexdigest()[:10]
    return f"{prefix}_{h}"


def ensure_dir(path: str | os.PathLike[str]) -> Path:
    p = Path(path)
    p.mkdir(parents=True, exist_ok=True)
    return p


def output_png_path(output_dir: str | os.PathLike[str], key: str, match_index: int, match_count: int) -> Path:
    out_dir = ensure_dir(output_dir)
    if match_count <= 1:
        name = f"{key}.png"
    else:
        name = f"{key}_{match_index}.png"
    return out_dir / name


def configure_logging(script_name: str) -> None:
    root = logging.getLogger()
    if root.handlers:
        return
    logging.basicConfig(
        level=logging.INFO,
        format=f"%(asctime)s [{script_name}] %(levelname)s: %(message)s",
    )


@dataclass(frozen=True)
class RunResult:
    required_key: str
    matched_files: list[Path]
    outputs: list[Path]
    failures: list[tuple[Path, str]]


def run_single_chart_script(
    *,
    data_dir: str | os.PathLike[str],
    output_dir: str | os.PathLike[str],
    required_key: str,
    render_one: Callable[[str | os.PathLike[str], str | os.PathLike[str]], None],
) -> RunResult:
    matched = scan_xlsx_by_key(data_dir, required_key)
    outputs: list[Path] = []
    failures: list[tuple[Path, str]] = []

    if not matched:
        msg = f"no xlsx matched required_key={required_key!r}"
        raise FileNotFoundError(msg)

    out_dir = ensure_dir(output_dir)
    logging.info("matched %d file(s) for key=%r in %s", len(matched), required_key, Path(data_dir))

    for i, xlsx_path in enumerate(matched, start=1):
        out_path = output_png_path(out_dir, required_key, i, len(matched))
        try:
            render_one(xlsx_path, out_path)
            outputs.append(out_path)
            logging.info("success: %s -> %s", xlsx_path.name, out_path.name)
        except Exception as e:  # pragma: no cover (covered by failure tests only)
            failures.append((xlsx_path, str(e)))
            logging.exception("failed: %s -> %s", xlsx_path.name, out_path.name)

    return RunResult(
        required_key=required_key,
        matched_files=matched,
        outputs=outputs,
        failures=failures,
    )


# ----------------------------
# Auto chart (rich template)
# ----------------------------

TIME_COL_CANDIDATES: list[str] = [
    "周起始日",
    "周起",
    "日期",
    "月份",
    "月",
    "年周",
    "year",
    "week_start",
]

CATEGORY_COL_CANDIDATES: list[str] = [
    "类型",
    "角色",
    "插件",
    "发布名",
    "评语任务",
    "问卷",
    "学部",
    "年级名称",
    "name",
    "学校",
]


def _to_number(v: object) -> float | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "":
        return None
    try:
        return float(s.replace(",", ""))
    except Exception:
        return None


def _to_time_key(v: object) -> str:
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v).strip()
    if "T" in s:
        s = s.split("T", 1)[0]
    return s


def infer_columns(headers: Sequence[str]) -> tuple[str | None, str | None, list[str]]:
    hset = [str(h).strip() for h in headers if h is not None and str(h).strip() != ""]
    time_col = next((c for c in TIME_COL_CANDIDATES if c in hset), None)
    cat_col = next((c for c in CATEGORY_COL_CANDIDATES if c in hset and c != time_col), None)

    # Numeric columns: all except time/category; later we check actual values.
    excluded = {time_col, cat_col, "周止", "周结束(周日)", "week", "year"}
    value_cols = [c for c in hset if c not in excluded]
    return time_col, cat_col, value_cols


def auto_render_rich_chart(
    *,
    xlsx_path: str | os.PathLike[str],
    out_path: str | os.PathLike[str],
    title: str | None = None,
    top_n: int = 5,
    include_total: bool = True,
) -> None:
    wb, ws = load_result_sheet(xlsx_path)
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {str(h).strip(): i + 1 for i, h in enumerate(headers) if h is not None and str(h).strip() != ""}

    time_col, cat_col, candidate_value_cols = infer_columns(list(idx.keys()))
    if time_col is None:
        # Fallback: treat first column as time-like.
        time_col = list(idx.keys())[0] if idx else None
    if time_col is None:
        raise RuntimeError("cannot infer time column")

    # Pick numeric columns by sampling.
    numeric_cols: list[str] = []
    for c in candidate_value_cols:
        if c == time_col or c == cat_col:
            continue
        col_i = idx.get(c)
        if col_i is None:
            continue
        ok = 0
        for r in range(2, min(ws.max_row, 50) + 1):
            if _to_number(ws.cell(r, col_i).value) is not None:
                ok += 1
                if ok >= 3:
                    break
        if ok >= 1:
            numeric_cols.append(c)

    if not numeric_cols:
        # As a last resort, allow 1st non-time column as numeric.
        for c in candidate_value_cols:
            if c != time_col and c != cat_col:
                numeric_cols = [c]
                break

    if not numeric_cols:
        raise RuntimeError("cannot infer numeric column")

    main_value_col = numeric_cols[0]

    time_i = idx[time_col]
    val_i = idx[main_value_col]
    cat_i = idx.get(cat_col) if cat_col else None

    # Aggregate
    by_time_cat: dict[str, dict[str, float]] = {}
    cat_total: dict[str, float] = {}
    by_time_total: dict[str, float] = {}

    for r in range(2, ws.max_row + 1):
        tv = ws.cell(r, time_i).value
        if tv is None or str(tv).strip() == "":
            continue
        tkey = _to_time_key(tv)

        num = _to_number(ws.cell(r, val_i).value)
        if num is None:
            continue

        if cat_i is not None:
            cv = ws.cell(r, cat_i).value
            cat = str(cv).strip() if cv is not None and str(cv).strip() != "" else "未知"
        else:
            cat = "合计"

        by_time_cat.setdefault(tkey, {})
        by_time_cat[tkey][cat] = by_time_cat[tkey].get(cat, 0.0) + float(num)
        cat_total[cat] = cat_total.get(cat, 0.0) + float(num)
        by_time_total[tkey] = by_time_total.get(tkey, 0.0) + float(num)

    if not by_time_cat:
        raise RuntimeError("no rows parsed")

    times = sorted(by_time_total.keys())
    x = list(range(len(times)))

    init_fonts()
    fig, ax = plt.subplots(figsize=(14, 6))

    if cat_i is not None:
        top_cats = [c for c, _ in sorted(cat_total.items(), key=lambda kv: kv[1], reverse=True)[:top_n]]
        # Include '其他' bucket for non-top categories.
        series: dict[str, list[float]] = {c: [] for c in top_cats}
        other: list[float] = []
        totals: list[float] = []

        for t in times:
            m = by_time_cat.get(t, {})
            total = float(by_time_total.get(t, 0.0))
            totals.append(total)
            s_other = 0.0
            for c, v in m.items():
                if c in series:
                    continue
                s_other += float(v)
            other.append(s_other)
            for c in top_cats:
                series[c].append(float(m.get(c, 0.0)))

        bottom = [0.0] * len(times)
        palette = ["#2563EB", "#10B981", "#F59E0B", "#8B5CF6", "#06B6D4"]
        for i, c in enumerate(top_cats):
            y = series[c]
            if max(y) == 0:
                continue
            ax.bar(x, y, bottom=bottom, label=c, color=palette[i % len(palette)], width=0.8)
            bottom = [b + v for b, v in zip(bottom, y)]

        if max(other) > 0:
            ax.bar(x, other, bottom=bottom, label="其他", color="#9CA3AF", width=0.8)

        if include_total:
            ax2 = ax.twinx()
            ax2.plot(x, totals, label="合计", color="#111827", linewidth=2.2, linestyle="--")
            ax2.set_ylabel("合计")

            h1, l1 = ax.get_legend_handles_labels()
            h2, l2 = ax2.get_legend_handles_labels()
            ax.legend(h1 + h2, l1 + l2, frameon=False, ncol=3, loc="upper left")
        else:
            ax.legend(frameon=False, loc="upper left")
    else:
        y = [float(by_time_total[t]) for t in times]
        ax.plot(x, y, label=main_value_col, color="#2563EB", linewidth=2.2)
        ax.legend(frameon=False)

    ax.set_xticks(x)
    ax.set_xticklabels(times, rotation=30, ha="right")
    ax.set_xlabel(time_col)
    ax.set_ylabel(main_value_col)
    ax.grid(True, axis="y", alpha=0.25)

    if title is None:
        title = f"{Path(xlsx_path).stem}"
    ax.set_title(title)

    fig.tight_layout()
    fig.savefig(out_path, dpi=180)
    plt.close(fig)

